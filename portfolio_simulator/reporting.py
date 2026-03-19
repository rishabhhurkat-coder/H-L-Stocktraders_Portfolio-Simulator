from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
import os
from pathlib import Path
import tempfile
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rich.text import Text

from portfolio_simulator.formatting import (
    format_currency,
    format_month_year,
    format_percentage,
    format_tenure,
)
from portfolio_simulator.models import Scenario, SimulationResult
from portfolio_simulator.simulation import add_months, sip_end_date, swp_start_date


@dataclass
class SheetPreview:
    name: str
    rows: list[tuple[str, str]] | None = None
    headers: list[str] | None = None
    table_rows: list[list[Any]] | None = None
    section_before_rows: list[int] | None = None
    section_after_rows: list[int] | None = None
    footer_rows: list[tuple[str, str]] | None = None
    footer_title: str | None = None


def build_export_previews(scenario: Scenario, result: SimulationResult) -> list[SheetPreview]:
    return [
        build_dashboard_sheet(scenario, result),
        build_cash_flow_sheet(result),
    ]


def export_reports(output_dir: Path, scenario: Scenario, result: SimulationResult) -> tuple[Path, Path]:
    previews = build_export_previews(scenario, result)
    workbook_path = output_dir / "MutualFundPortfolioSimulator.xlsx"
    csv_path = output_dir / "MutualFundPortfolioSimulator_CashFlow.csv"

    export_workbook(workbook_path, previews)
    export_cash_flow_csv(csv_path, next(sheet for sheet in previews if sheet.name == "Cash Flow"))

    return workbook_path, csv_path


PDF_MODE_ORDER = [
    "All Combination Mode",
    "SIP Only Mode",
    "SIP + Investment CF Mode",
    "SIP + SWP Mode",
]


def excel_schedule_columns_for_mode(mode_label: str) -> list[str]:
    if mode_label == "SIP Only Mode":
        return ["#", "Date", "Type", "SIP", "Net CF", "Growth", "Close"]
    if mode_label == "SIP + Investment CF Mode":
        return ["#", "Date", "Type", "SIP", "Investment CF", "Net CF", "Growth", "Close"]
    if mode_label == "SIP + SWP Mode":
        return ["#", "Date", "Type", "SIP", "SWP", "Net CF", "Growth", "Close"]
    return ["#", "Date", "Type", "SIP", "SWP", "Investment CF", "Net CF", "Growth", "Close"]


def export_excel_report(
    output_dir: Path,
    scenario: Scenario,
    analysis_results: dict[str, SimulationResult],
    selected_variant: str,
    customer_profile: dict[str, Any] | None = None,
    developer_name: str = "Rishabh Hurkat",
    developer_phone: str = "88830488312",
    developer_email: str = "hlstocktraders@gmail.com",
) -> tuple[str, bytes]:
    _ = output_dir
    workbook_name = "MutualFundPortfolioSimulator_FullReport.xlsx"

    workbook = Workbook()
    workbook.remove(workbook.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    title_font = Font(bold=True, color="103B52", size=13)
    label_font = Font(bold=True, color="103B52")
    note_font = Font(color="2F3E4E", italic=True)

    def currency_text(amount: float) -> str:
        return format_currency(amount)

    def goal_based_invested(result: SimulationResult) -> float:
        return float(sum((row.sip_amount - row.swp_amount + row.lumpsum_amount) for row in result.schedule_rows))

    def mode_metrics(result: SimulationResult) -> dict[str, float]:
        return {
            "Total Invested": goal_based_invested(result),
            "Final Portfolio": float(result.final_portfolio_value),
            "Total Profit": float(result.total_profit),
        }

    def selected_or_fallback() -> SimulationResult:
        if selected_variant in analysis_results:
            return analysis_results[selected_variant]
        if "All Combination Mode" in analysis_results:
            return analysis_results["All Combination Mode"]
        return next(iter(analysis_results.values()))

    def write_title(sheet: Any, title: str) -> int:
        sheet.cell(row=1, column=1, value=title).font = title_font
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        return 3

    def write_table(sheet: Any, start_row: int, headers: list[str], rows: list[list[str]]) -> int:
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(row=start_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        row_ptr = start_row + 1
        for row_values in rows:
            for col_idx, value in enumerate(row_values, start=1):
                sheet.cell(row=row_ptr, column=col_idx, value=value)
            row_ptr += 1
        return row_ptr - 1

    def write_notes(sheet: Any, start_row: int, analysis_note: str, recommendation: str) -> int:
        sheet.cell(row=start_row, column=1, value=f"Analysis Note: {analysis_note}").font = note_font
        sheet.cell(row=start_row + 1, column=1, value=f"Recommendation: {recommendation}").font = note_font
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=8)
        sheet.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=8)
        return start_row + 2

    def sanitize_sheet_name(name: str) -> str:
        cleaned = (
            name.replace("/", " ")
            .replace("\\", " ")
            .replace("*", " ")
            .replace("[", " ")
            .replace("]", " ")
            .replace(":", " ")
            .replace("?", " ")
        ).strip()
        return cleaned[:31]

    current_result = selected_or_fallback()
    metrics_map = {mode: mode_metrics(result) for mode, result in analysis_results.items()}
    mode_order = [mode for mode in PDF_MODE_ORDER if mode in analysis_results]

    # Client and header details
    profile = customer_profile or {}
    sheet = workbook.create_sheet("Client & Header")
    row_start = write_title(sheet, "Mutual Fund Portfolio Simulator - Full Report")
    profile_headers = ["Field", "Value"]
    profile_rows = [
        ["Developer Name", developer_name],
        ["Developer Contact", developer_phone],
        ["Developer Email", developer_email],
        ["Customer Name", str(profile.get("name", "-"))],
        ["Birth Date", profile.get("birth_date").strftime("%d-%m-%Y") if profile.get("birth_date") else "-"],
        ["Occupation", str(profile.get("occupation", "-"))],
        ["Address", str(profile.get("address", "-"))],
        ["City", str(profile.get("city", "-"))],
        ["Contact Details", str(profile.get("contact_details", "-"))],
        ["Generated At", datetime.now().strftime("%d-%m-%Y %H:%M")],
    ]
    end_row = write_table(sheet, row_start, profile_headers, profile_rows)
    write_notes(
        sheet,
        end_row + 2,
        "Customer and advisor context used for the current analysis export.",
        "Confirm customer profile and communication details before sharing.",
    )

    # Scenario inputs
    sheet = workbook.create_sheet("Scenario Inputs")
    row_start = write_title(sheet, "Scenario Inputs")
    scenario_rows = [
        ["SIP Start Date", format_month_year(scenario.sip_start_date)],
        ["SIP End Date", format_month_year(sip_end_date(scenario))],
        ["Investment Tenure", format_tenure(scenario.investment_years, scenario.investment_months)],
        ["Monthly SIP", currency_text(scenario.monthly_sip)],
        ["Expected Return", format_percentage(scenario.annual_roi)],
        ["Inflation", format_percentage(scenario.inflation_rate)],
        ["Step-Up", f"{scenario.step_up_rate:.2f} %" if scenario.step_up_enabled else "No"],
        ["Selected Variant", selected_variant],
        ["Reference Final Portfolio", currency_text(current_result.final_portfolio_value)],
    ]
    end_row = write_table(sheet, row_start, ["Field", "Value"], scenario_rows)
    write_notes(
        sheet,
        end_row + 2,
        "Assumptions in this section drive all result calculations.",
        "Use conservative return assumptions and realistic cash flow expectations.",
    )

    # Investment cash flows
    sheet = workbook.create_sheet("Investment CF")
    row_start = write_title(sheet, "Investment Cash Flows")
    cf_rows: list[list[str]] = []
    if scenario.cash_flow_events:
        for event in sorted(scenario.cash_flow_events, key=lambda item: item.event_date):
            signed = event.amount if event.flow_type == "add" else -event.amount
            cf_rows.append(
                [
                    event.flow_type.title(),
                    event.event_date.strftime("%d-%m-%Y"),
                    currency_text(event.amount),
                    currency_text(signed),
                ]
            )
    else:
        cf_rows.append(["-", "-", "-", "-"])
    end_row = write_table(sheet, row_start, ["Type", "Date", "Amount", "Signed Impact"], cf_rows)
    write_notes(
        sheet,
        end_row + 2,
        "Cash flow additions/withdrawals alter deployed capital and net-invested track.",
        "Capture major life-event cash flows to avoid over/under-estimated projections.",
    )

    # SWP details
    sheet = workbook.create_sheet("SWP Details")
    row_start = write_title(sheet, "SWP Details")
    swp_rows: list[list[str]] = []
    if scenario.swp_enabled:
        swp_month_count = max(1, scenario.swp_years * 12)
        swp_rows.extend(
            [
                ["SWP Status", "Enabled"],
                ["SWP Start", format_month_year(swp_start_date(scenario))],
                ["SWP End", format_month_year(add_months(swp_start_date(scenario) or scenario.sip_start_date, swp_month_count - 1))],
                ["SWP Tenure", format_tenure(scenario.swp_years, scenario.swp_months)],
                ["Monthly SWP", currency_text(current_result.actual_monthly_swp)],
                ["Total Withdrawal", currency_text(current_result.actual_monthly_swp * swp_month_count)],
                ["Remaining Fund", currency_text(current_result.final_portfolio_value)],
            ]
        )
    else:
        swp_rows.extend(
            [
                ["SWP Status", "Not Enabled"],
                ["Default End (SIP)", format_month_year(sip_end_date(scenario))],
            ]
        )
    end_row = write_table(sheet, row_start, ["Field", "Value"], swp_rows)
    write_notes(
        sheet,
        end_row + 2,
        "SWP sustainability depends on amount, tenure and sequence of returns.",
        "Keep SWP within sustainable limit and review annually.",
    )

    # Simulation results
    sheet = workbook.create_sheet("Simulation Results")
    row_start = write_title(sheet, "Simulation Results - Mode Comparison")
    sim_rows: list[list[str]] = []
    sip_reference = metrics_map.get("SIP Only Mode", {}).get("Final Portfolio", 0.0)
    for mode in mode_order:
        metrics = metrics_map[mode]
        delta = "-" if mode == "SIP Only Mode" else currency_text(metrics["Final Portfolio"] - sip_reference)
        sim_rows.append(
            [
                mode,
                currency_text(metrics["Total Invested"]),
                currency_text(metrics["Final Portfolio"]),
                currency_text(metrics["Total Profit"]),
                delta,
            ]
        )
    end_row = write_table(
        sheet,
        row_start,
        ["Mode", "Total Invested", "Final Portfolio", "Total Profit", "Change vs SIP Only"],
        sim_rows,
    )
    write_notes(
        sheet,
        end_row + 2,
        "Comparative output highlights impact of cash flows and SWP strategy choices.",
        "Select mode with strong final value and acceptable withdrawal flexibility.",
    )

    # Comparison schedule
    sheet = workbook.create_sheet("Comparison Schedule")
    row_start = write_title(sheet, "Final Comparison Schedule")
    comp_rows: list[list[str]] = []
    headers = ["Metric", *mode_order]
    for metric in ("Final Portfolio", "Total Invested", "Total Profit"):
        comp_rows.append([metric, *[currency_text(metrics_map[mode][metric]) for mode in mode_order]])
    change_values = ["Change vs SIP Only"]
    for mode in mode_order:
        if mode == "SIP Only Mode":
            change_values.append("-")
        else:
            change_values.append(currency_text(metrics_map[mode]["Final Portfolio"] - sip_reference))
    comp_rows.append(change_values)
    end_row = write_table(sheet, row_start, headers, comp_rows)
    write_notes(
        sheet,
        end_row + 2,
        "Comparison schedule summarizes mode-wise outcomes in one place.",
        "Use SIP Only baseline and adopt advanced mode only when goal need is clear.",
    )

    # Mode-wise schedule sheets
    for mode in mode_order:
        result = analysis_results[mode]
        headers = excel_schedule_columns_for_mode(mode)
        schedule_rows: list[list[str]] = []
        for row in result.schedule_rows:
            if row.period_date is None:
                continue
            net_cf = row.sip_amount - row.swp_amount + row.lumpsum_amount
            row_map = {
                "#": str(row.period_number),
                "Date": row.period_date.strftime("%b %Y"),
                "Type": row.phase,
                "SIP": currency_text(row.sip_amount),
                "SWP": currency_text(row.swp_amount),
                "Investment CF": currency_text(row.lumpsum_amount),
                "Net CF": currency_text(net_cf),
                "Growth": currency_text(row.growth),
                "Close": currency_text(row.closing_balance),
            }
            schedule_rows.append([row_map[column] for column in headers])
        if not schedule_rows:
            schedule_rows.append(["-"] * len(headers))

        sheet = workbook.create_sheet(sanitize_sheet_name(f"CF {mode}"))
        row_start = write_title(sheet, f"Cash Flow Schedule - {mode}")
        end_row = write_table(sheet, row_start, headers, schedule_rows)

        write_notes(
            sheet,
            end_row + 2,
            "Schedule tracks monthly cash movement, growth and closing portfolio.",
            "Review periods with low cushion and rebalance SIP/SWP/cash-flow timing.",
        )

    for worksheet in workbook.worksheets:
        autosize_columns(worksheet)

    output_buffer = BytesIO()
    workbook.save(output_buffer)
    return workbook_name, output_buffer.getvalue()


def export_pdf_report(
    output_dir: Path,
    scenario: Scenario,
    analysis_results: dict[str, SimulationResult],
    selected_variant: str,
    customer_profile: dict[str, Any] | None = None,
    logo_path: Path | None = None,
    developer_name: str = "Rishabh Hurkat",
    developer_phone: str = "88830488312",
    developer_email: str = "hlstocktraders@gmail.com",
) -> Path:
    try:
        from reportlab.graphics.shapes import Drawing, Line, PolyLine, Rect, String
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import (
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ModuleNotFoundError as exc:
        raise RuntimeError(
            "PDF export requires reportlab. Install dependencies from requirements.txt and retry."
        ) from exc

    output_path = output_dir / "MutualFundPortfolioSimulator_FullReport.pdf"

    font_regular = "Helvetica"
    font_bold = "Helvetica-Bold"
    font_italic = "Helvetica-Oblique"
    rupee_supported = False

    font_candidates: list[tuple[str, Path, Path, Path]] = []
    windir = os.environ.get("WINDIR", "")
    if windir:
        font_dir = Path(windir) / "Fonts"
        font_candidates.extend(
            [
                ("PDFArial", font_dir / "arial.ttf", font_dir / "arialbd.ttf", font_dir / "ariali.ttf"),
                ("PDFSegoeUI", font_dir / "segoeui.ttf", font_dir / "segoeuib.ttf", font_dir / "segoeuii.ttf"),
            ]
        )

    font_candidates.extend(
        [
            (
                "PDFDejaVuSans",
                Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
                Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
                Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Oblique.ttf"),
            ),
            (
                "PDFLiberationSans",
                Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf"),
                Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf"),
                Path("/usr/share/fonts/truetype/liberation2/LiberationSans-Italic.ttf"),
            ),
            (
                "PDFNotoSans",
                Path("/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf"),
                Path("/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf"),
                Path("/usr/share/fonts/truetype/noto/NotoSans-Italic.ttf"),
            ),
        ]
    )

    for font_name, regular_path, bold_path, italic_path in font_candidates:
        if not (regular_path.exists() and bold_path.exists() and italic_path.exists()):
            continue
        try:
            pdfmetrics.registerFont(TTFont(font_name, str(regular_path)))
            pdfmetrics.registerFont(TTFont(f"{font_name}Bold", str(bold_path)))
            pdfmetrics.registerFont(TTFont(f"{font_name}Italic", str(italic_path)))
            font_regular = font_name
            font_bold = f"{font_name}Bold"
            font_italic = f"{font_name}Italic"
            rupee_supported = True
            break
        except Exception:
            continue

    doc = SimpleDocTemplate(
        str(output_path),
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=34 * mm,
        bottomMargin=14 * mm,
        title="Mutual Fund Portfolio Simulator - Full Report",
        author="H&L Stock Traders",
    )

    styles = getSampleStyleSheet()
    section_style = ParagraphStyle(
        "SectionTitle",
        parent=styles["Heading2"],
        fontName=font_bold,
        fontSize=13,
        textColor=colors.HexColor("#0B7285"),
        spaceBefore=8,
        spaceAfter=6,
    )
    body_style = ParagraphStyle(
        "Body",
        parent=styles["BodyText"],
        fontName=font_regular,
        fontSize=9.5,
        leading=13,
        textColor=colors.HexColor("#16202A"),
    )
    note_style = ParagraphStyle(
        "Note",
        parent=styles["BodyText"],
        fontName=font_italic,
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#2F3E4E"),
        backColor=colors.HexColor("#F4F8FC"),
        borderPadding=6,
        borderWidth=0.4,
        borderColor=colors.HexColor("#D7E2EE"),
        spaceBefore=4,
        spaceAfter=8,
    )
    small_style = ParagraphStyle(
        "Small",
        parent=styles["BodyText"],
        fontName=font_regular,
        fontSize=8.4,
        leading=11,
        textColor=colors.HexColor("#4D5B6B"),
    )

    def safe_paragraph(value: str) -> str:
        return (
            str(value)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace("\n", "<br/>")
        )

    def goal_based_invested(result: SimulationResult) -> float:
        return float(sum((row.sip_amount - row.swp_amount + row.lumpsum_amount) for row in result.schedule_rows))

    def currency_text(amount: float) -> str:
        text = format_currency(amount)
        if rupee_supported:
            return text
        return text.replace("\u20b9", "INR")

    def selected_or_fallback() -> SimulationResult:
        if selected_variant in analysis_results:
            return analysis_results[selected_variant]
        if "All Combination Mode" in analysis_results:
            return analysis_results["All Combination Mode"]
        return next(iter(analysis_results.values()))

    def mode_metrics(result: SimulationResult) -> dict[str, float]:
        return {
            "Total Invested": goal_based_invested(result),
            "Final Portfolio": float(result.final_portfolio_value),
            "Total Profit": float(result.total_profit),
        }

    def comparison_rows(metrics_map: dict[str, dict[str, float]]) -> list[list[str]]:
        modes = [mode for mode in PDF_MODE_ORDER if mode in metrics_map]
        rows: list[list[str]] = [["Metric", *modes]]
        for metric in ("Final Portfolio", "Total Invested", "Total Profit"):
            rows.append([metric, *[currency_text(metrics_map[mode][metric]) for mode in modes]])
        sip_reference = metrics_map.get("SIP Only Mode", {}).get("Final Portfolio", 0.0)
        delta_row = ["Change vs SIP Only"]
        for mode in modes:
            if mode == "SIP Only Mode":
                delta_row.append("-")
            else:
                delta_row.append(currency_text(metrics_map[mode]["Final Portfolio"] - sip_reference))
        rows.append(delta_row)
        return rows

    def build_table(data: list[list[str]], col_widths: list[float] | None = None, repeat_rows: int = 1) -> Table:
        table = Table(data, colWidths=col_widths, repeatRows=repeat_rows)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAF1F8")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#123B55")),
                    ("FONTNAME", (0, 0), (-1, 0), font_bold),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                    ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CFDAE6")),
                    ("FONTNAME", (0, 1), (-1, -1), font_regular),
                    ("FONTSIZE", (0, 1), (-1, -1), 8.6),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        return table

    def add_segment_note(story: list[Any], text: str) -> None:
        story.append(Paragraph(f"<b>Analysis Note:</b> {safe_paragraph(text)}", note_style))

    def build_comparison_chart(metrics_map: dict[str, dict[str, float]]) -> Drawing:
        modes = [mode for mode in PDF_MODE_ORDER if mode in metrics_map]
        categories = ["Final Portfolio", "Total Invested", "Total Profit"]
        series_colors = ["#3A86FF", "#2EC4B6", "#FF9F1C", "#FF006E"]

        drawing = Drawing(520, 260)
        chart_x, chart_y = 42, 52
        chart_w, chart_h = 440, 160
        drawing.add(Line(chart_x, chart_y, chart_x, chart_y + chart_h, strokeColor=colors.HexColor("#9FB3C8")))
        drawing.add(Line(chart_x, chart_y, chart_x + chart_w, chart_y, strokeColor=colors.HexColor("#9FB3C8")))

        values = [metrics_map[mode][cat] for mode in modes for cat in categories]
        min_v = min([0.0, *values]) if values else 0.0
        max_v = max([1.0, *values]) if values else 1.0
        if max_v == min_v:
            max_v += 1.0
        range_v = max_v - min_v

        def y_scale(value: float) -> float:
            return chart_y + ((value - min_v) / range_v) * chart_h

        zero_y = y_scale(0.0)
        drawing.add(Line(chart_x, zero_y, chart_x + chart_w, zero_y, strokeColor=colors.HexColor("#D5E0EB")))

        group_w = chart_w / max(len(categories), 1)
        bar_w = group_w / (max(len(modes), 1) + 0.8)

        for cat_idx, category in enumerate(categories):
            base_x = chart_x + (cat_idx * group_w)
            for mode_idx, mode in enumerate(modes):
                value = metrics_map[mode][category]
                y_val = y_scale(value)
                rect_y = min(zero_y, y_val)
                rect_h = max(1.2, abs(y_val - zero_y))
                rect_x = base_x + (mode_idx * bar_w) + 10
                drawing.add(
                    Rect(
                        rect_x,
                        rect_y,
                        bar_w * 0.75,
                        rect_h,
                        fillColor=colors.HexColor(series_colors[mode_idx % len(series_colors)]),
                        strokeColor=colors.HexColor(series_colors[mode_idx % len(series_colors)]),
                    )
                )
            drawing.add(
                String(
                    base_x + (group_w * 0.5) - 28,
                    chart_y - 14,
                    category,
                    fontName=font_regular,
                    fontSize=7.7,
                    fillColor=colors.HexColor("#435364"),
                )
            )

        legend_x = chart_x
        legend_y = chart_y + chart_h + 24
        for idx, mode in enumerate(modes):
            lx = legend_x + (idx * 120)
            drawing.add(Rect(lx, legend_y, 8, 8, fillColor=colors.HexColor(series_colors[idx % len(series_colors)]), strokeColor=None))
            drawing.add(
                String(
                    lx + 12,
                    legend_y - 1,
                    mode.replace(" Mode", ""),
                    fontName=font_regular,
                    fontSize=7.5,
                    fillColor=colors.HexColor("#2A3A4C"),
                )
            )

        return drawing

    def build_growth_chart(result: SimulationResult, mode_label: str) -> Drawing:
        rows = [row for row in result.schedule_rows if row.period_date is not None]
        drawing = Drawing(520, 230)
        chart_x, chart_y = 42, 42
        chart_w, chart_h = 440, 142
        drawing.add(Line(chart_x, chart_y, chart_x, chart_y + chart_h, strokeColor=colors.HexColor("#9FB3C8")))
        drawing.add(Line(chart_x, chart_y, chart_x + chart_w, chart_y, strokeColor=colors.HexColor("#9FB3C8")))
        drawing.add(
            String(
                chart_x,
                chart_y + chart_h + 28,
                f"Portfolio Growth - {mode_label}",
                fontName=font_bold,
                fontSize=10,
                fillColor=colors.HexColor("#123B55"),
            )
        )

        if not rows:
            drawing.add(
                String(
                    chart_x + 120,
                    chart_y + 60,
                    "No timeline data available.",
                    fontName=font_regular,
                    fontSize=8.5,
                    fillColor=colors.HexColor("#6C7B89"),
                )
            )
            return drawing

        portfolio = [float(row.closing_balance) for row in rows]
        net_series: list[float] = []
        rolling = 0.0
        for row in rows:
            rolling += float(row.sip_amount - row.swp_amount + row.lumpsum_amount)
            net_series.append(rolling)

        min_v = min([0.0, *portfolio, *net_series])
        max_v = max([1.0, *portfolio, *net_series])
        if max_v == min_v:
            max_v += 1.0
        range_v = max_v - min_v

        def x_scale(idx: int) -> float:
            if len(rows) == 1:
                return chart_x
            return chart_x + (idx * chart_w / (len(rows) - 1))

        def y_scale(value: float) -> float:
            return chart_y + ((value - min_v) / range_v) * chart_h

        zero_y = y_scale(0.0)
        drawing.add(Line(chart_x, zero_y, chart_x + chart_w, zero_y, strokeColor=colors.HexColor("#D5E0EB")))

        portfolio_points: list[float] = []
        net_points: list[float] = []
        for idx, value in enumerate(portfolio):
            portfolio_points.extend([x_scale(idx), y_scale(value)])
        for idx, value in enumerate(net_series):
            net_points.extend([x_scale(idx), y_scale(value)])

        drawing.add(PolyLine(portfolio_points, strokeColor=colors.HexColor("#1D4ED8"), strokeWidth=1.8))
        drawing.add(PolyLine(net_points, strokeColor=colors.HexColor("#FFD60A"), strokeWidth=1.7))

        start_date = rows[0].period_date.strftime("%b %Y") if rows[0].period_date else "-"
        end_date = rows[-1].period_date.strftime("%b %Y") if rows[-1].period_date else "-"
        drawing.add(String(chart_x, chart_y - 14, start_date, fontName=font_regular, fontSize=7.5, fillColor=colors.HexColor("#5B6A79")))
        drawing.add(
            String(
                chart_x + chart_w - 38,
                chart_y - 14,
                end_date,
                fontName=font_regular,
                fontSize=7.5,
                fillColor=colors.HexColor("#5B6A79"),
            )
        )

        legend_y = chart_y + chart_h + 8
        drawing.add(Rect(chart_x, legend_y, 10, 3, fillColor=colors.HexColor("#FFD60A"), strokeColor=None))
        drawing.add(String(chart_x + 14, legend_y - 2, "Cumulative Net Cash Flow", fontName=font_regular, fontSize=7.5))
        drawing.add(Rect(chart_x + 170, legend_y, 10, 3, fillColor=colors.HexColor("#1D4ED8"), strokeColor=None))
        drawing.add(String(chart_x + 184, legend_y - 2, "Portfolio Value", fontName=font_regular, fontSize=7.5))
        return drawing

    def cash_flow_event_rows() -> list[list[str]]:
        rows = [["Type", "Date", "Amount", "Signed Impact"]]
        if not scenario.cash_flow_events:
            rows.append(["-", "-", "-", "-"])
            return rows
        for event in sorted(scenario.cash_flow_events, key=lambda item: item.event_date):
            signed = event.amount if event.flow_type == "add" else -event.amount
            rows.append(
                [
                    event.flow_type.title(),
                    event.event_date.strftime("%d-%m-%Y"),
                    currency_text(event.amount),
                    currency_text(signed),
                ]
            )
        return rows

    def swp_detail_rows(reference_result: SimulationResult) -> list[list[str]]:
        rows = [["Field", "Value"]]
        if not scenario.swp_enabled:
            rows.extend(
                [
                    ["SWP Status", "Not Enabled"],
                    ["Default End (SIP)", format_month_year(sip_end_date(scenario))],
                ]
            )
            return rows

        swp_month_count = max(1, scenario.swp_years * 12)
        rows.extend(
            [
                ["SWP Status", "Enabled"],
                ["SWP Start", format_month_year(swp_start_date(scenario))],
                ["SWP End", format_month_year(add_months(swp_start_date(scenario) or scenario.sip_start_date, swp_month_count - 1))],
                ["SWP Tenure", format_tenure(scenario.swp_years, scenario.swp_months)],
                ["Monthly SWP", currency_text(reference_result.actual_monthly_swp)],
                ["Total Withdrawal", currency_text(reference_result.actual_monthly_swp * swp_month_count)],
                ["Remaining Fund", currency_text(reference_result.final_portfolio_value)],
            ]
        )
        return rows

    def scenario_detail_rows() -> list[list[str]]:
        selected_result = selected_or_fallback()
        rows = [
            ["Field", "Value"],
            ["SIP Start Date", format_month_year(scenario.sip_start_date)],
            ["SIP End Date", format_month_year(sip_end_date(scenario))],
            ["Investment Tenure", format_tenure(scenario.investment_years, scenario.investment_months)],
            ["Monthly SIP", currency_text(scenario.monthly_sip)],
            ["Expected Return", format_percentage(scenario.annual_roi)],
            ["Inflation", format_percentage(scenario.inflation_rate)],
            ["Step-Up", f"{scenario.step_up_rate:.2f} %" if scenario.step_up_enabled else "No"],
            ["Cash Flows Enabled", "Yes" if bool(scenario.cash_flow_events) else "No"],
            ["SWP Enabled", "Yes" if scenario.swp_enabled else "No"],
            ["Selected Variant", selected_variant],
            ["Reference Final Portfolio", currency_text(selected_result.final_portfolio_value)],
        ]
        return rows

    def simulation_result_rows(metrics_map: dict[str, dict[str, float]]) -> list[list[str]]:
        rows = [["Mode", "Total Invested", "Final Portfolio", "Total Profit"]]
        for mode in PDF_MODE_ORDER:
            if mode not in metrics_map:
                continue
            metrics = metrics_map[mode]
            rows.append(
                [
                    mode,
                    currency_text(metrics["Total Invested"]),
                    currency_text(metrics["Final Portfolio"]),
                    currency_text(metrics["Total Profit"]),
                ]
            )
        return rows

    def schedule_columns_for_mode(mode_label: str) -> list[str]:
        if mode_label == "SIP Only Mode":
            return ["#", "Date", "Type", "SIP", "Net CF", "Growth", "Close"]
        if mode_label == "SIP + Investment CF Mode":
            return ["#", "Date", "Type", "SIP", "Investment CF", "Net CF", "Growth", "Close"]
        if mode_label == "SIP + SWP Mode":
            return ["#", "Date", "Type", "SIP", "SWP", "Net CF", "Growth", "Close"]
        return ["#", "Date", "Type", "SIP", "SWP", "Investment CF", "Net CF", "Growth", "Close"]

    def schedule_rows_for_mode(mode_label: str, result: SimulationResult) -> list[list[str]]:
        selected_columns = schedule_columns_for_mode(mode_label)
        rows = [selected_columns]
        for row in result.schedule_rows:
            if row.period_date is None:
                continue
            net_cf = row.sip_amount - row.swp_amount + row.lumpsum_amount
            row_map = {
                "#": str(row.period_number),
                "Date": row.period_date.strftime("%b %Y"),
                "Type": row.phase,
                "SIP": currency_text(row.sip_amount),
                "SWP": currency_text(row.swp_amount),
                "Investment CF": currency_text(row.lumpsum_amount),
                "Net CF": currency_text(net_cf),
                "Growth": currency_text(row.growth),
                "Close": currency_text(row.closing_balance),
            }
            rows.append([row_map[column] for column in selected_columns])
        if len(rows) == 1:
            rows.append(["-"] * len(selected_columns))
        return rows

    logo_reader = None
    if logo_path is not None and logo_path.exists():
        try:
            logo_reader = ImageReader(str(logo_path))
        except Exception:
            logo_reader = None

    def draw_header(canvas: Any, doc_obj: Any) -> None:
        canvas.saveState()
        page_w, page_h = A4
        left = doc_obj.leftMargin
        right = page_w - doc_obj.rightMargin
        header_top = page_h - 10 * mm
        header_bottom = page_h - 28 * mm

        canvas.setStrokeColor(colors.HexColor("#CFDAE6"))
        canvas.setLineWidth(0.7)
        canvas.line(left, header_bottom, right, header_bottom)

        if logo_reader is not None:
            canvas.drawImage(
                logo_reader,
                left,
                header_bottom + 1.5 * mm,
                width=18 * mm,
                height=15 * mm,
                preserveAspectRatio=True,
                mask="auto",
            )

        text_x = left + 22 * mm
        canvas.setFont(font_bold, 10)
        canvas.setFillColor(colors.HexColor("#103B52"))
        canvas.drawString(text_x, header_top - 6 * mm, "Mutual Fund Portfolio Simulator - H&L Stock Traders")
        canvas.setFont(font_regular, 8.4)
        canvas.setFillColor(colors.HexColor("#3D4E60"))
        canvas.drawString(
            text_x,
            header_top - 11 * mm,
            f"{developer_name}  |  {developer_phone}  |  {developer_email}",
        )
        canvas.setFont(font_regular, 8)
        canvas.drawRightString(right, header_top - 6 * mm, f"Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}")
        canvas.drawRightString(right, header_top - 11 * mm, f"Page {canvas.getPageNumber()}")
        canvas.restoreState()

    story: list[Any] = []
    current_result = selected_or_fallback()
    metrics_map = {mode: mode_metrics(result) for mode, result in analysis_results.items()}
    comparison_table_rows = comparison_rows(metrics_map)

    story.append(Paragraph("Mutual Fund Portfolio Simulator - Full Analysis Report", styles["Title"]))
    story.append(Spacer(1, 3))
    story.append(Paragraph("Client Details", section_style))
    profile = customer_profile or {}
    profile_table = [
        ["Field", "Value"],
        ["Customer Name", str(profile.get("name", "-"))],
        ["Birth Date", profile.get("birth_date").strftime("%d-%m-%Y") if profile.get("birth_date") else "-"],
        ["Occupation", str(profile.get("occupation", "-"))],
        ["Address", str(profile.get("address", "-"))],
        ["City", str(profile.get("city", "-"))],
        ["Contact Details", str(profile.get("contact_details", "-"))],
    ]
    story.append(build_table(profile_table, col_widths=[60 * mm, 120 * mm]))
    add_segment_note(
        story,
        "This report uses current assumptions and captured customer details. Recommendation: validate customer details and risk profile before final submission.",
    )

    story.append(Paragraph("Scenario Inputs", section_style))
    story.append(build_table(scenario_detail_rows(), col_widths=[60 * mm, 120 * mm]))
    add_segment_note(
        story,
        "SIP tenure, return, inflation, cash flow and SWP inputs directly drive all outputs. Recommendation: keep assumptions conservative and align with client goal horizon.",
    )

    story.append(Paragraph("Investment Cash Flows", section_style))
    story.append(build_table(cash_flow_event_rows(), col_widths=[30 * mm, 35 * mm, 40 * mm, 45 * mm]))
    add_segment_note(
        story,
        "Cash flow adds increase deployed capital and withdrawals reduce net invested base. Recommendation: map every major expected add/withdraw event for realistic planning.",
    )

    story.append(Paragraph("SWP Details", section_style))
    story.append(build_table(swp_detail_rows(current_result), col_widths=[60 * mm, 120 * mm]))
    add_segment_note(
        story,
        "SWP sustainability depends on withdrawal amount, start timing and return behavior. Recommendation: keep SWP below sustainable level to reduce depletion risk.",
    )

    story.append(PageBreak())
    story.append(Paragraph("Simulation Results - All Modes", section_style))
    story.append(build_table(simulation_result_rows(metrics_map), col_widths=[58 * mm, 40 * mm, 45 * mm, 45 * mm]))
    best_mode = max(
        (mode for mode in PDF_MODE_ORDER if mode in metrics_map),
        key=lambda mode: metrics_map[mode]["Final Portfolio"],
    ) if metrics_map else "-"
    add_segment_note(
        story,
        f"Comparative analysis indicates best final portfolio in {best_mode}. Recommendation: prioritize the mode balancing higher final value with acceptable cash flow flexibility.",
    )

    story.append(Paragraph("Scenario Comparison Chart", section_style))
    story.append(build_comparison_chart(metrics_map))
    add_segment_note(
        story,
        "Bar chart highlights impact differences across modes for invested amount, final value and profit. Recommendation: evaluate both upside and stability, not only peak return.",
    )

    story.append(PageBreak())
    story.append(Paragraph("Portfolio Growth Charts (Mode-wise)", section_style))
    for mode in PDF_MODE_ORDER:
        if mode not in analysis_results:
            continue
        story.append(build_growth_chart(analysis_results[mode], mode))
        story.append(Spacer(1, 4))
    add_segment_note(
        story,
        "Growth lines compare cumulative net cash flow against portfolio value through time. Recommendation: prefer trajectories with stable gap expansion and manageable drawdown periods.",
    )

    for mode in PDF_MODE_ORDER:
        if mode not in analysis_results:
            continue
        story.append(PageBreak())
        story.append(Paragraph(f"Cash Flow Schedule - {mode}", section_style))
        schedule_rows = schedule_rows_for_mode(mode, analysis_results[mode])
        col_count = len(schedule_rows[0]) if schedule_rows else 1
        available_width = 180 * mm
        dynamic_col_widths = [available_width / col_count] * col_count
        story.append(
            build_table(
                schedule_rows,
                col_widths=dynamic_col_widths,
            )
        )
        add_segment_note(
            story,
            f"Schedule for {mode} shows month-wise contributions, withdrawals, growth and closing values. Recommendation: review periods with low growth cushion before approving execution.",
        )

    story.append(PageBreak())
    story.append(Paragraph("Final Comparison Schedule", section_style))
    story.append(build_table(comparison_table_rows, col_widths=[45 * mm, 36 * mm, 36 * mm, 36 * mm, 36 * mm]))
    story.append(Paragraph("Overall Recommendation", section_style))
    story.append(
        Paragraph(
            safe_paragraph(
                "Use SIP Only as conservative baseline and compare with All Combination for goal acceleration. "
                "When withdrawals are expected, prefer calibrated SIP + SWP or SIP + CF with periodic reviews. "
                "Revalidate assumptions quarterly and after major life or market changes."
            ),
            body_style,
        )
    )
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            safe_paragraph("Prepared for advisory discussion. This is a scenario-based analytical output, not a guaranteed return commitment."),
            small_style,
        )
    )

    doc.build(story, onFirstPage=draw_header, onLaterPages=draw_header)
    return output_path


def export_pdf_report_bytes(
    scenario: Scenario,
    analysis_results: dict[str, SimulationResult],
    selected_variant: str,
    customer_profile: dict[str, Any] | None = None,
    logo_path: Path | None = None,
    developer_name: str = "Rishabh Hurkat",
    developer_phone: str = "88830488312",
    developer_email: str = "hlstocktraders@gmail.com",
) -> tuple[str, bytes]:
    with tempfile.TemporaryDirectory() as temp_dir:
        pdf_path = export_pdf_report(
            Path(temp_dir),
            scenario,
            analysis_results,
            selected_variant,
            customer_profile=customer_profile,
            logo_path=logo_path,
            developer_name=developer_name,
            developer_phone=developer_phone,
            developer_email=developer_email,
        )
        return pdf_path.name, pdf_path.read_bytes()


def export_workbook(path: Path, previews: list[SheetPreview]) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    for preview in previews:
        sheet = workbook.create_sheet(title=preview.name[:31])
        row_index = 1

        if preview.rows:
            for label, value in preview.rows:
                sheet.cell(row=row_index, column=1, value=label)
                sheet.cell(row=row_index, column=2, value=value)
                sheet.cell(row=row_index, column=1).font = Font(bold=True, color="00B7EB")
                row_index += 1
            row_index += 1

        if preview.headers:
            for col_index, header in enumerate(preview.headers, start=1):
                cell = sheet.cell(row=row_index, column=col_index, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
            row_index += 1

        if preview.table_rows:
            for table_row in preview.table_rows:
                for col_index, value in enumerate(table_row, start=1):
                    sheet.cell(row=row_index, column=col_index, value=cell_text(value))
                row_index += 1

        if preview.footer_rows:
            row_index += 1
            title_cell = sheet.cell(row=row_index, column=1, value=preview.footer_title or "SUMMARY")
            title_cell.font = Font(bold=True, color="00B7EB")
            row_index += 1
            for label, value in preview.footer_rows:
                sheet.cell(row=row_index, column=1, value=label)
                sheet.cell(row=row_index, column=2, value=value)
                sheet.cell(row=row_index, column=1).font = Font(bold=True, color="00B7EB")
                row_index += 1

        autosize_columns(sheet)

    workbook.save(path)


def export_cash_flow_csv(path: Path, preview: SheetPreview) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        if preview.headers:
            writer.writerow(preview.headers)
        for row in preview.table_rows or []:
            writer.writerow([cell_text(value) for value in row])
        if preview.footer_rows:
            writer.writerow([])
            writer.writerow([preview.footer_title or "SUMMARY"])
            for label, value in preview.footer_rows:
                writer.writerow([label, value])


def cell_text(value: Any) -> str:
    if isinstance(value, Text):
        return value.plain
    return str(value)


def autosize_columns(sheet: Any) -> None:
    for idx, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        first_cell = column_cells[0]
        column_letter = getattr(first_cell, "column_letter", get_column_letter(idx))
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)


def build_dashboard_sheet(scenario: Scenario, result: SimulationResult) -> SheetPreview:
    total_horizon_months = result.total_months_simulated
    total_horizon_years = total_horizon_months / 12 if total_horizon_months else 0
    cash_flow_text = "No"
    if scenario.cash_flow_events:
        adds = sum(event.amount for event in scenario.cash_flow_events if event.flow_type == "add")
        withdrawals = sum(event.amount for event in scenario.cash_flow_events if event.flow_type == "withdraw")
        cash_flow_text = (
            f"{len(scenario.cash_flow_events)} items / "
            f"Add {format_currency(adds)} / Withdraw {format_currency(withdrawals)}"
        )

    swp_text = "No"
    if scenario.swp_enabled:
        swp_text = (
            f"{format_month_year(swp_start_date(scenario))} / "
            f"{format_tenure(scenario.swp_years, scenario.swp_months)} / "
            f"{format_currency(result.actual_monthly_swp)}"
        )

    rows = [
        ("Dashboard", "Scenario + Outcome Summary"),
        ("SIP Start Date", format_month_year(scenario.sip_start_date)),
        ("SIP End Date", format_month_year(sip_end_date(scenario))),
        ("Investment Tenure", format_tenure(scenario.investment_years, scenario.investment_months)),
        ("Monthly SIP", format_currency(scenario.monthly_sip)),
        ("Expected Return", format_percentage(scenario.annual_roi)),
        ("Inflation", format_percentage(scenario.inflation_rate)),
        ("Step-Up SIP", f"{scenario.step_up_rate:.2f} % yearly" if scenario.step_up_enabled else "No"),
        ("Cash Flows", cash_flow_text),
        ("SWP Start Date", format_month_year(swp_start_date(scenario))),
        ("SWP Plan", swp_text),
        ("Total Horizon", f"{total_horizon_years:.2f} Years" if total_horizon_years else "--"),
        ("Total Invested", format_currency(result.total_invested)),
        ("SIP Phase Value", format_currency(result.sip_end_value)),
        ("SWP Start Value", format_currency(result.swp_start_value)),
        ("Monthly SWP Used", format_currency(result.actual_monthly_swp)),
        ("Final Portfolio Value", format_currency(result.final_portfolio_value)),
        ("Total Profit", format_currency(result.total_profit)),
        ("Portfolio CAGR", format_percentage(result.cagr * 100)),
        ("Inflation Adjusted Value", format_currency(result.inflation_adjusted_value)),
    ]
    return SheetPreview(name="Dashboard", rows=rows)


def build_scenario_sheet(scenario: Scenario, result: SimulationResult) -> SheetPreview:
    cash_flow_text = "No"
    if scenario.cash_flow_events:
        adds = sum(event.amount for event in scenario.cash_flow_events if event.flow_type == "add")
        withdrawals = sum(event.amount for event in scenario.cash_flow_events if event.flow_type == "withdraw")
        cash_flow_text = (
            f"{len(scenario.cash_flow_events)} items / "
            f"Add {format_currency(adds)} / Withdraw {format_currency(withdrawals)}"
        )

    swp_text = "No"
    if scenario.swp_enabled:
        swp_text = (
            f"{format_month_year(swp_start_date(scenario))} / "
            f"{format_tenure(scenario.swp_years, scenario.swp_months)} / "
            f"{format_currency(result.actual_monthly_swp)}"
        )

    rows = [
        ("SIP Start Date", format_month_year(scenario.sip_start_date)),
        ("Monthly SIP", format_currency(scenario.monthly_sip)),
        ("Investment Tenure", format_tenure(scenario.investment_years, scenario.investment_months)),
        ("Expected Return", format_percentage(scenario.annual_roi)),
        ("Inflation", format_percentage(scenario.inflation_rate)),
        ("Step-Up SIP", f"{scenario.step_up_rate:.2f} % yearly" if scenario.step_up_enabled else "No"),
        ("Cash Flows", cash_flow_text),
        ("SWP Plan", swp_text),
    ]
    return SheetPreview(name="Scenario Summary", rows=rows)


def build_cash_flow_sheet(result: SimulationResult) -> SheetPreview:
    headers = ["#", "Date", "Type", "SIP", "SWP", "LumpSum", "Net CF", "Growth", "Value (Close)"]
    table_rows = []
    section_before_rows: list[int] = []
    section_after_rows: list[int] = []
    year_sip_total = 0.0
    year_swp_total = 0.0
    year_lumpsum_total = 0.0
    year_net_total = 0.0
    year_growth_total = 0.0
    current_year: int | None = None
    total_sip = 0.0
    total_swp = 0.0
    total_lumpsum = 0.0
    total_net = 0.0
    total_growth = 0.0

    for index, row in enumerate(result.schedule_rows, start=1):
        row_year = row.period_date.year if row.period_date else None
        if current_year is None:
            current_year = row_year
        elif row_year != current_year:
            section_before_rows.append(len(table_rows))
            table_rows.append(
                year_total_row(
                    current_year,
                    year_sip_total,
                    year_swp_total,
                    year_lumpsum_total,
                    year_net_total,
                    year_growth_total,
                    result.schedule_rows[index - 2].closing_balance,
                )
            )
            section_after_rows.append(len(table_rows) - 1)
            year_sip_total = 0.0
            year_swp_total = 0.0
            year_lumpsum_total = 0.0
            year_net_total = 0.0
            year_growth_total = 0.0
            current_year = row_year

        net_cf = row.sip_amount - row.swp_amount + row.lumpsum_amount
        lumpsum_style = amount_style(row.lumpsum_amount)
        year_sip_total += row.sip_amount
        year_swp_total += row.swp_amount
        year_lumpsum_total += row.lumpsum_amount
        year_net_total += net_cf
        year_growth_total += row.growth
        total_sip += row.sip_amount
        total_swp += row.swp_amount
        total_lumpsum += row.lumpsum_amount
        total_net += net_cf
        total_growth += row.growth
        table_rows.append(
            [
                str(index),
                format_month_year(row.period_date),
                row.phase,
                styled_currency(row.sip_amount, "green"),
                styled_currency(row.swp_amount, "red"),
                styled_currency(row.lumpsum_amount, lumpsum_style),
                styled_currency(net_cf, amount_style(net_cf)),
                format_currency(row.growth),
                format_currency(row.closing_balance),
            ]
        )
    if result.schedule_rows and current_year is not None:
        section_before_rows.append(len(table_rows))
        table_rows.append(
            year_total_row(
                current_year,
                year_sip_total,
                year_swp_total,
                year_lumpsum_total,
                year_net_total,
                year_growth_total,
                result.schedule_rows[-1].closing_balance,
            )
        )
        section_after_rows.append(len(table_rows) - 1)

    footer_rows = [
        ("Total SIP", format_currency(total_sip)),
        ("Total SWP", format_currency(total_swp)),
        ("Total LumpSum", format_currency(total_lumpsum)),
        ("Total Net CF", format_currency(total_net)),
        ("Total Growth", format_currency(total_growth)),
        ("Final Value", format_currency(result.final_portfolio_value)),
    ]

    return SheetPreview(
        name="Cash Flow",
        headers=headers,
        table_rows=table_rows,
        section_before_rows=section_before_rows,
        section_after_rows=section_after_rows,
        footer_rows=footer_rows,
        footer_title="GRAND TOTAL",
    )


def amount_style(amount: float) -> str:
    if amount > 0:
        return "green"
    if amount < 0:
        return "red"
    return ""


def styled_currency(amount: float, style: str) -> Text:
    return Text(format_currency(amount), style=style)


def year_total_row(
    year: int,
    sip_total: float,
    swp_total: float,
    lumpsum_total: float,
    net_total: float,
    growth_total: float,
    closing_balance: float,
) -> list[Any]:
    return [
        "",
        str(year),
        Text("Year Total", style="bold cyan"),
        styled_currency(sip_total, "green"),
        styled_currency(swp_total, "red"),
        styled_currency(lumpsum_total, amount_style(lumpsum_total)),
        styled_currency(net_total, amount_style(net_total)),
        Text(format_currency(growth_total), style="bold cyan"),
        Text(format_currency(closing_balance), style="bold cyan"),
    ]


def build_sip_schedule_sheet(result: SimulationResult) -> SheetPreview:
    headers = ["#", "Date", "SIP", "LumpSum", "Growth", "Close"]
    table_rows = [
        [
            str(row.period_number),
            format_month_year(row.period_date),
            format_currency(row.sip_amount),
            format_currency(row.lumpsum_amount),
            format_currency(row.growth),
            format_currency(row.closing_balance),
        ]
        for row in result.schedule_rows
        if row.sip_amount > 0 or row.lumpsum_amount != 0
    ]
    return SheetPreview(name="SIP Schedule", headers=headers, table_rows=table_rows)


def build_swp_schedule_sheet(result: SimulationResult) -> SheetPreview:
    headers = ["#", "Date", "SWP", "Growth", "Close"]
    table_rows = [
        [
            str(row.period_number),
            format_month_year(row.period_date),
            format_currency(row.swp_amount),
            format_currency(row.growth),
            format_currency(row.closing_balance),
        ]
        for row in result.schedule_rows
        if row.swp_amount > 0
    ]
    return SheetPreview(name="SWP Schedule", headers=headers, table_rows=table_rows)
